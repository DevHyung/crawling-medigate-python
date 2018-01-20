#-*-encoding:utf8-*-
import urllib.request
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Alignment, Font,Border,Side,NamedStyle,colors
import re
numregex = re.compile('\d+')
###
font =Font(color=colors.BLACK)
font2 = Font(color='578fcc')
font3 = Font(color="ff0000")
fill = PatternFill("solid",  fgColor='FF839ce3')
ali = Alignment(horizontal='center',vertical='center',shrinkToFit=True)
thin = Side(border_style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
###
def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)
    rows = ws[cell_range]

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom
    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill
                c.font = font
                c.alignment=alignment
                c.border = border
def initExcel(filename):
    header1 = ['지역','분류','과목','기관','기관명','제목','조회','작성일','마감']
    wb = Workbook()
    ws1 = wb.worksheets[0]
    ws1.append(header1)
    style_range(ws1, 'A1:I1', border=border, fill=fill,font=font, alignment=ali)
    wb.save(filename)
def saveExcel(datalist,imagelist,alist,filename):
    wb = load_workbook(filename)
    ws1 = wb.worksheets[0]
    startrow = ws1.max_row + 1
    ##size
    date_style = NamedStyle(name='datetime', number_format='MM월 DD일')
    for rowidx in range(len(datalist)):
        idx = 0
        for td in datalist[rowidx]:

            if idx==0 or idx == 4:
                ws1.cell(row=startrow, column=idx + 1, value=td).font = font2
            elif idx == 5:
                ws1.cell(row=startrow, column=idx + 1).value = '=HYPERLINK("{}", "{}")'.format(alist[rowidx], td)
                ws1.cell(row=startrow, column=idx + 1).font = font2
            elif idx==6:
                ws1.cell(row=startrow, column=idx + 1, value=int(td))
            elif idx == 7:
                mm,dd=td.split('-')
                str1 = mm+"월 "+dd+"일"
                ws1.cell(row=startrow, column=idx + 1, value=str1)
            elif idx == 8:
                if 'D-' in td or '오늘' in td:
                    ws1.cell(row=startrow, column=idx + 1, value=td).font = font3
                else:
                    mm, dd = td.split('-')
                    str1 = mm + "월 " + dd + "일"
                    ws1.cell(row=startrow, column=idx + 1, value=str1)
            else:
                ws1.cell(row=startrow, column=idx + 1, value=td)
            idx += 1
        startrow += 1
    colidx = 0
    for col in ws1.columns:
        column = col[0].column  # Get the column name
        if colidx == 4:
            ws1.column_dimensions[column].width = 20
        else:
            ws1.column_dimensions[column].width = 10
        colidx+=1
    wb.save(filename)

if __name__=="__main__":
    print("_______________________________________________")
    print("*** Made By Pakr HyungJune copyright @ DevHyung")
    f = open("option.txt", 'r',encoding='utf8')
    option = f.readlines()
    ID = option[2].strip()
    PW = option[6].strip()
    FILENAME = option[10].strip()
    startpage = int(option[14].strip())
    endpage = int(option[18].strip())
    cycle = int(option[22].strip())
    detailURL = 'http://www.medigate.net/cbiz/recjob.do?cmd=detail&menuGroupCode=CBIZ&menuCode=RECJOB&pageNo=1&inviteType=&hopLocCode=&hopCityCode=&spcCode=&orgType=&searchKey=title&searchValue=&ctgCode=job&boardIdx='
    initExcel(FILENAME)
    print("_______________________________________________")

    driver = webdriver.Chrome('./chromedriver')
    driver.get('http://www.medigate.net/index.jsp')
    # login start
    driver.find_element_by_xpath('//*[@id="contentID"]/div[1]/div[2]/div[1]/form/fieldset/div[2]/input[1]').send_keys(ID)
    driver.find_element_by_xpath('//*[@id="contentID"]/div[1]/div[2]/div[1]/form/fieldset/div[2]/input[2]').send_keys(PW)
    driver.find_element_by_xpath('//*[@id="contentID"]/div[1]/div[2]/div[1]/form/fieldset/div[2]/button').click()
    time.sleep(0.5)
    picidx = 2
    try:
        for pagenum in range(startpage,endpage+1):
            print(">>>",pagenum," 페이지 추출중")
            driver.get('http://www.medigate.net/cbiz/recjob.do?cmd=list&menuGroupCode=CBIZ&menuCode=RECJOB&ctgCode=job&pageNo='+str(pagenum))
            bs4 = BeautifulSoup(driver.page_source,'lxml')
            table = bs4.find_all('div',class_="wrap_board_list")[2].find('table')
            trlist = table.find_all('tr')[1:]
            datalist = []
            alist = []
            imagelist = []
            for tr in trlist:
                tdlist = tr.find_all('td')
                tmplist = []
                for td in tdlist:
                    try:
                        tmplist.append(td.find('a').get_text().strip())
                        alist.append(detailURL+str(numregex.search(td.find('a')['href']).group()))
                    except:
                        tmplist.append(td.get_text().strip())
                datalist.append(tmplist)
            alistidx = 0
            print("\t>>> 엑셀에 저장중...")
            for detail in alist:
                driver.get(detail)
                time.sleep(cycle)
                bs4 = BeautifulSoup(driver.page_source,'lxml')
                try:
                    img = bs4.find('div',class_="wrap_head").find('img')['src']
                    urllib.request.urlretrieve(img,"./image/"+str(picidx)+"_"+datalist[alistidx][4]+".jpg")
                    imagelist.append("./image/"+detail.split('boardIdx=')[1]+".jpg")
                except:
                    pass
                picidx+=1
                alistidx+=1
            saveExcel(datalist,imagelist,alist,FILENAME)
            print("\t>>> 엑셀에 저장완료...")
            print(">>>", pagenum, " 페이지 추출완료 !")
    except:
        print("일시적으로 네트워크 불안정, 타겟사이트 불안정으로 오류가 날수있습니다.")
        print("계속해서 오류가 나면 연락주세요 ㅡ 박형준")
        saveExcel(datalist, imagelist, alist, FILENAME)
    driver.quit()
